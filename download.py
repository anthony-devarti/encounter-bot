# download.py
from __future__ import annotations

from io import BytesIO
from typing import Optional, List, Tuple

import aiosqlite
from openpyxl import Workbook


async def _fetch_table_def(
    db: aiosqlite.Connection,
    guild_id: int,
    group_key: str,
    region_id: Optional[int],
    type_key: Optional[str],
) -> Optional[aiosqlite.Row]:
    cur = await db.execute(
        """
        SELECT id, roll_mode
        FROM table_def
        WHERE guild_id = ?
          AND group_key = ?
          AND region_id IS ?
          AND type_key IS ?
        """,
        (guild_id, group_key, region_id, type_key),
    )
    return await cur.fetchone()


async def _fetch_table_entries(db: aiosqlite.Connection, table_id: int) -> List[aiosqlite.Row]:
    cur = await db.execute(
        """
        SELECT min_roll, max_roll, weight, result
        FROM table_entry
        WHERE table_id = ?
        ORDER BY sort_order ASC
        """,
        (table_id,),
    )
    return await cur.fetchall()


def _add_sheet_for_table(
    wb: Workbook,
    title: str,
    kind: str,       # "type" or "result"
    roll_mode: str,  # uniform/weight/range
    entries: List[aiosqlite.Row],
) -> None:
    ws = wb.create_sheet(title=title)

    if roll_mode == "range":
        headers = ["min", "max", kind]
    elif roll_mode == "weight":
        headers = ["weight", kind]
    else:
        headers = [kind]

    ws.append(headers)

    for e in entries:
        if roll_mode == "range":
            ws.append([e["min_roll"], e["max_roll"], e["result"]])
        elif roll_mode == "weight":
            ws.append([e["weight"], e["result"]])
        else:
            ws.append([e["result"]])


async def _fetch_regions(db: aiosqlite.Connection, guild_id: int) -> List[Tuple[int, str]]:
    """
    Returns regions in configured order (sort_order).
    """
    cur = await db.execute(
        """
        SELECT region_id, region_name
        FROM region
        WHERE guild_id = ?
        ORDER BY sort_order ASC, region_id ASC
        """,
        (guild_id,),
    )
    rows = await cur.fetchall()
    return [(int(r["region_id"]), r["region_name"]) for r in rows]


def _enc_type_tab(region_id: Optional[int]) -> str:
    return "Encounter Types" if region_id is None else f"Encounter Types - {region_id}"


def _enc_tab(region_id: Optional[int], t: str) -> str:
    return f"Encounter - {t}" if region_id is None else f"Encounter - {region_id} - {t}"


def _rew_tab(region_id: Optional[int], t: str) -> str:
    return f"Reward - {t}" if region_id is None else f"Reward - {region_id} - {t}"


async def build_workbook_bytes(db: aiosqlite.Connection, guild_id: int) -> bytes:
    """
    Exports the currently imported tables into an XLSX workbook.

    Two possible exports:
      - Default mode (no regions configured): Encounter Types + Encounter/Reward - <Type>
      - Regional mode: Regions sheet + Encounter Types - <region_id> + Encounter/Reward - <region_id> - <Type>
    """
    wb = Workbook()
    wb.remove(wb.active)

    regions = await _fetch_regions(db, guild_id)
    regional_mode = len(regions) > 0

    if regional_mode:
        # Regions sheet
        ws = wb.create_sheet("Regions")
        ws.append(["region_id", "region_name"])
        for rid, name in regions:
            ws.append([rid, name])

        # For each region, export Encounter Types - <rid>, then typed tables
        for rid, _name in regions:
            enc_type_def = await _fetch_table_def(db, guild_id, "encounter_type", rid, None)
            if not enc_type_def:
                # skip region if no data (shouldn't happen if imported properly)
                continue

            enc_type_entries = await _fetch_table_entries(db, enc_type_def["id"])
            _add_sheet_for_table(
                wb,
                title=_enc_type_tab(rid),
                kind="type",
                roll_mode=enc_type_def["roll_mode"],
                entries=enc_type_entries,
            )

            # types in stored order
            types = [e["result"] for e in enc_type_entries if e["result"]]

            for t in types:
                enc_def = await _fetch_table_def(db, guild_id, "encounter", rid, t)
                if enc_def:
                    enc_entries = await _fetch_table_entries(db, enc_def["id"])
                    _add_sheet_for_table(
                        wb,
                        title=_enc_tab(rid, t),
                        kind="result",
                        roll_mode=enc_def["roll_mode"],
                        entries=enc_entries,
                    )

                rew_def = await _fetch_table_def(db, guild_id, "reward", rid, t)
                if rew_def:
                    rew_entries = await _fetch_table_entries(db, rew_def["id"])
                    _add_sheet_for_table(
                        wb,
                        title=_rew_tab(rid, t),
                        kind="result",
                        roll_mode=rew_def["roll_mode"],
                        entries=rew_entries,
                    )
    else:
        # Default mode
        enc_type_def = await _fetch_table_def(db, guild_id, "encounter_type", None, None)
        if not enc_type_def:
            raise RuntimeError("No imported data found. Run /import first.")

        enc_type_entries = await _fetch_table_entries(db, enc_type_def["id"])
        _add_sheet_for_table(
            wb,
            title=_enc_type_tab(None),
            kind="type",
            roll_mode=enc_type_def["roll_mode"],
            entries=enc_type_entries,
        )

        types = [e["result"] for e in enc_type_entries if e["result"]]

        for t in types:
            enc_def = await _fetch_table_def(db, guild_id, "encounter", None, t)
            if enc_def:
                enc_entries = await _fetch_table_entries(db, enc_def["id"])
                _add_sheet_for_table(
                    wb,
                    title=_enc_tab(None, t),
                    kind="result",
                    roll_mode=enc_def["roll_mode"],
                    entries=enc_entries,
                )

            rew_def = await _fetch_table_def(db, guild_id, "reward", None, t)
            if rew_def:
                rew_entries = await _fetch_table_entries(db, rew_def["id"])
                _add_sheet_for_table(
                    wb,
                    title=_rew_tab(None, t),
                    kind="result",
                    roll_mode=rew_def["roll_mode"],
                    entries=rew_entries,
                )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
