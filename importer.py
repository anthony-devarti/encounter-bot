# importer.py
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import aiosqlite
from openpyxl import load_workbook

ROLL_UNIFORM = "uniform"
ROLL_WEIGHT = "weight"
ROLL_RANGE = "range"


@dataclass
class ImportCounts:
    encounter_types: int
    encounter_entries: int
    reward_types: int
    reward_entries: int
    regions: int


@dataclass
class ImportError:
    tab: str
    row: Optional[int]
    message: str


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _norm_header(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        if v.is_integer():
            return int(v)
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        f = float(s)
        if f.is_integer():
            return int(f)
        return None
    except Exception:
        return None


def _read_sheet_rows(ws) -> Tuple[Dict[str, int], List[Tuple[int, List[Any]]]]:
    """
    Returns:
      header_map: normalized header -> column index
      data: list of (excel_row_number, row_values)
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, []

    header = rows[0]
    header_map: Dict[str, int] = {}
    for idx, h in enumerate(header):
        key = _norm_header(h)
        if key and key not in header_map:
            header_map[key] = idx

    data: List[Tuple[int, List[Any]]] = []
    for i, r in enumerate(rows[1:], start=2):
        if r is None:
            continue
        if all((c is None or str(c).strip() == "") for c in r):
            continue
        data.append((i, list(r)))

    return header_map, data


def _detect_mode(tab: str, header_map: Dict[str, int], data: List[Tuple[int, List[Any]]]) -> Tuple[str, List[ImportError]]:
    """
    Mode precedence:
      - if any min/max columns exist, treat as range (requires both)
      - else if weight exists, treat as weight
      - else uniform
    """
    errors: List[ImportError] = []

    has_min = "min" in header_map
    has_max = "max" in header_map
    has_weight = "weight" in header_map

    if has_min or has_max:
        if not (has_min and has_max):
            errors.append(ImportError(tab, None, "Range mode requires both 'min' and 'max' columns."))
            return ROLL_RANGE, errors

        for row_no, row in data:
            mi = _to_int(row[header_map["min"]])
            ma = _to_int(row[header_map["max"]])
            if mi is None or ma is None:
                continue
            if mi > ma:
                errors.append(ImportError(tab, row_no, f"Invalid range: min {mi} is greater than max {ma}."))
        return ROLL_RANGE, errors

    if has_weight:
        for row_no, row in data:
            w = _to_int(row[header_map["weight"]])
            if w is None:
                continue
            if w <= 0:
                errors.append(ImportError(tab, row_no, f"Invalid weight {w}. Must be a positive integer."))
        return ROLL_WEIGHT, errors

    return ROLL_UNIFORM, errors


def _validate_ranges(tab: str, header_map: Dict[str, int], data: List[Tuple[int, List[Any]]]) -> List[ImportError]:
    errors: List[ImportError] = []
    ranges: List[Tuple[int, int, int]] = []

    for row_no, row in data:
        mi = _to_int(row[header_map["min"]])
        ma = _to_int(row[header_map["max"]])
        if mi is None or ma is None:
            continue
        ranges.append((mi, ma, row_no))

    ranges.sort(key=lambda x: (x[0], x[1]))
    for i in range(1, len(ranges)):
        prev_min, prev_max, prev_row = ranges[i - 1]
        cur_min, cur_max, cur_row = ranges[i]
        if cur_min <= prev_max:
            errors.append(
                ImportError(
                    tab,
                    cur_row,
                    f"Overlapping ranges with row {prev_row}: {prev_min}-{prev_max} overlaps {cur_min}-{cur_max}.",
                )
            )
    return errors


def _parse_type_tab(wb, tab: str) -> Tuple[str, List[str], Optional[int], List[ImportError]]:
    """
    Reads a type table tab.
    Expected column: 'type'
    Optional columns: ('min','max') or ('weight')
    Returns: (roll_mode, types, max_roll, errors)
    """
    errors: List[ImportError] = []

    if tab not in wb.sheetnames:
        return ROLL_UNIFORM, [], None, [ImportError(tab, None, "Missing required tab.")]

    ws = wb[tab]
    header_map, data = _read_sheet_rows(ws)

    if "type" not in header_map:
        return ROLL_UNIFORM, [], None, [ImportError(tab, None, "Missing required column 'type'.")]

    mode, mode_errs = _detect_mode(tab, header_map, data)
    errors.extend(mode_errs)

    max_roll: Optional[int] = None
    types: List[str] = []

    if mode == ROLL_RANGE:
        errors.extend(_validate_ranges(tab, header_map, data))
        mr = 0
        for _, row in data:
            t = _cell_str(row[header_map["type"]])
            if not t:
                continue
            mi = _to_int(row[header_map["min"]])
            ma = _to_int(row[header_map["max"]])
            if mi is None or ma is None:
                continue
            mr = max(mr, ma)
            types.append(t)
        max_roll = mr if mr > 0 else None

    elif mode == ROLL_WEIGHT:
        for _, row in data:
            t = _cell_str(row[header_map["type"]])
            if not t:
                continue
            w = _to_int(row[header_map["weight"]])
            if w is None or w <= 0:
                continue
            types.append(t)

    else:
        for _, row in data:
            t = _cell_str(row[header_map["type"]])
            if t:
                types.append(t)

    # preserve order, remove duplicates
    types = list(dict.fromkeys(types))

    if not types:
        errors.append(ImportError(tab, None, "No types found."))

    return mode, types, max_roll, errors


def _parse_result_tab(wb, tab: str) -> Tuple[str, List[Dict[str, Any]], Optional[int], List[ImportError]]:
    """
    Reads an encounter or reward tab.
    Expected column: 'result'
    Optional columns: ('min','max') or ('weight')
    Returns: (roll_mode, entries, max_roll, errors)

    Each entry dict contains:
      - min (Optional[int])
      - max (Optional[int])
      - weight (Optional[int])
      - result (str)
    """
    errors: List[ImportError] = []

    if tab not in wb.sheetnames:
        return ROLL_UNIFORM, [], None, [ImportError(tab, None, "Missing required tab.")]

    ws = wb[tab]
    header_map, data = _read_sheet_rows(ws)

    if "result" not in header_map:
        return ROLL_UNIFORM, [], None, [ImportError(tab, None, "Missing required column 'result'.")]

    mode, mode_errs = _detect_mode(tab, header_map, data)
    errors.extend(mode_errs)

    entries: List[Dict[str, Any]] = []
    max_roll: Optional[int] = None

    if mode == ROLL_RANGE:
        errors.extend(_validate_ranges(tab, header_map, data))
        mr = 0
        for _, row in data:
            result = _cell_str(row[header_map["result"]])
            if not result:
                continue
            mi = _to_int(row[header_map["min"]])
            ma = _to_int(row[header_map["max"]])
            if mi is None or ma is None:
                continue
            mr = max(mr, ma)
            entries.append({"min": mi, "max": ma, "weight": None, "result": result})
        max_roll = mr if mr > 0 else None

    elif mode == ROLL_WEIGHT:
        for _, row in data:
            result = _cell_str(row[header_map["result"]])
            if not result:
                continue
            w = _to_int(row[header_map["weight"]])
            if w is None or w <= 0:
                continue
            entries.append({"min": None, "max": None, "weight": w, "result": result})

    else:
        for _, row in data:
            result = _cell_str(row[header_map["result"]])
            if result:
                entries.append({"min": None, "max": None, "weight": None, "result": result})

    if not entries:
        errors.append(ImportError(tab, None, "No results found."))

    return mode, entries, max_roll, errors


def _parse_regions_tab(wb) -> Tuple[List[Tuple[int, str]], List[ImportError]]:
    """
    Regions tab format:
      Tab: Regions
      Columns: region_id, region_name

    Returns:
      regions: list of (region_id, region_name) in sheet order
    """
    tab = "Regions"
    if tab not in wb.sheetnames:
        return [], []

    ws = wb[tab]
    header_map, data = _read_sheet_rows(ws)

    errs: List[ImportError] = []
    if "region_id" not in header_map or "region_name" not in header_map:
        errs.append(ImportError(tab, None, "Regions tab must have columns: region_id, region_name."))
        return [], errs

    regions: List[Tuple[int, str]] = []
    seen: set[int] = set()

    for row_no, row in data:
        rid = _to_int(row[header_map["region_id"]])
        name = _cell_str(row[header_map["region_name"]])

        if rid is None:
            errs.append(ImportError(tab, row_no, "region_id must be an integer."))
            continue
        if rid <= 0:
            errs.append(ImportError(tab, row_no, "region_id must be a positive integer."))
            continue
        if not name:
            errs.append(ImportError(tab, row_no, "region_name is required."))
            continue
        if rid in seen:
            errs.append(ImportError(tab, row_no, f"Duplicate region_id {rid}."))
            continue

        seen.add(rid)
        regions.append((rid, name))

    return regions, errs


def _enc_type_tab(region_id: Optional[int]) -> str:
    return "Encounter Types" if region_id is None else f"Encounter Types - {region_id}"


def _enc_tab(region_id: Optional[int], t: str) -> str:
    return f"Encounter - {t}" if region_id is None else f"Encounter - {region_id} - {t}"


def _rew_tab(region_id: Optional[int], t: str) -> str:
    return f"Reward - {t}" if region_id is None else f"Reward - {region_id} - {t}"


async def import_workbook_bytes(
    db: aiosqlite.Connection,
    guild_id: int,
    xlsx_bytes: bytes,
) -> Tuple[Optional[ImportCounts], List[ImportError]]:
    """
    Imports XLSX into SQLite.

    Supports two layouts:

    1) Default (no Regions tab):
       - Encounter Types
       - Encounter - <Type>
       - Reward - <Type>

    2) Regional (Regions tab present):
       - Regions (region_id, region_name)
       - Encounter Types - <region_id>
       - Encounter - <region_id> - <Type>
       - Reward - <region_id> - <Type>

    Reward type is not rolled separately. Reward tables are keyed off encounter type.
    A "Reward Types" tab may exist, but it is ignored.
    """
    wb = load_workbook(filename=BytesIO(xlsx_bytes), data_only=True)
    sheetnames = set(wb.sheetnames)

    errors: List[ImportError] = []

    regions, region_errs = _parse_regions_tab(wb)
    errors.extend(region_errs)

    # If Regions tab exists, we are in regional mode
    regional_mode = ("Regions" in sheetnames)

    # Validate regions presence and required tabs
    region_ids: List[Optional[int]]
    if regional_mode:
        if region_errs:
            return None, errors
        if not regions:
            return None, [ImportError("Regions", None, "Regions tab is present but has no valid rows.")]
        region_ids = [rid for rid, _ in regions]
    else:
        region_ids = [None]

    # Parse and validate all regions first, without touching the DB
    parsed_regions: List[Dict[str, Any]] = []
    total_enc_types = 0
    total_enc_entries = 0
    total_rew_entries = 0

    for region_id in region_ids:
        type_tab = _enc_type_tab(region_id)
        enc_type_mode, types, enc_type_max, type_errs = _parse_type_tab(wb, type_tab)
        errors.extend(type_errs)
        if type_errs:
            continue

        # Required per-type tabs
        for t in types:
            enc_tab = _enc_tab(region_id, t)
            rew_tab = _rew_tab(region_id, t)

            if enc_tab not in sheetnames:
                errors.append(ImportError(enc_tab, None, f"Missing tab for encounter type '{t}'."))
            if rew_tab not in sheetnames:
                errors.append(ImportError(rew_tab, None, f"Missing tab for reward type '{t}'."))

        if errors:
            continue

        enc_tables: Dict[str, Dict[str, Any]] = {}
        rew_tables: Dict[str, Dict[str, Any]] = {}

        for t in types:
            enc_tab = _enc_tab(region_id, t)
            mode, entries, max_roll, tab_errs = _parse_result_tab(wb, enc_tab)
            errors.extend(tab_errs)
            enc_tables[t] = {"tab": enc_tab, "mode": mode, "entries": entries, "max_roll": max_roll}

        for t in types:
            rew_tab = _rew_tab(region_id, t)
            mode, entries, max_roll, tab_errs = _parse_result_tab(wb, rew_tab)
            errors.extend(tab_errs)
            rew_tables[t] = {"tab": rew_tab, "mode": mode, "entries": entries, "max_roll": max_roll}

        if errors:
            continue

        parsed_regions.append(
            {
                "region_id": region_id,
                "enc_type": {"tab": type_tab, "mode": enc_type_mode, "max_roll": enc_type_max, "types": types},
                "enc_tables": enc_tables,
                "rew_tables": rew_tables,
            }
        )

        total_enc_types += len(types)
        total_enc_entries += sum(len(enc_tables[t]["entries"]) for t in types)
        total_rew_entries += sum(len(rew_tables[t]["entries"]) for t in types)

    if errors:
        return None, errors

    updated_at = _now_iso()

    await db.execute("BEGIN")
    try:
        await db.execute(
            """
            INSERT INTO guild_config(guild_id, updated_at) VALUES(?, ?)
            ON CONFLICT(guild_id) DO UPDATE SET updated_at=excluded.updated_at
            """,
            (guild_id, updated_at),
        )

        # Replace all tables and regions for this guild
        await db.execute("DELETE FROM table_def WHERE guild_id = ?", (guild_id,))
        await db.execute("DELETE FROM region WHERE guild_id = ?", (guild_id,))

        # Insert regions mapping (regional mode only)
        if regional_mode:
            for i, (rid, name) in enumerate(regions):
                await db.execute(
                    """
                    INSERT INTO region(guild_id, region_id, region_name, sort_order)
                    VALUES(?,?,?,?)
                    """,
                    (guild_id, rid, name, i),
                )

        async def insert_table(
            group_key: str,
            region_id: Optional[int],
            type_key: Optional[str],
            roll_mode: str,
            max_roll: Optional[int],
        ) -> int:
            cur = await db.execute(
                """
                INSERT INTO table_def(guild_id, group_key, region_id, type_key, roll_mode, max_roll, updated_at)
                VALUES(?,?,?,?,?,?,?)
                """,
                (guild_id, group_key, region_id, type_key, roll_mode, max_roll, updated_at),
            )
            return cur.lastrowid

        async def insert_entries(table_id: int, entries: List[Dict[str, Any]]) -> None:
            for i, e in enumerate(entries):
                await db.execute(
                    """
                    INSERT INTO table_entry(table_id, min_roll, max_roll, weight, result, sort_order)
                    VALUES(?,?,?,?,?,?)
                    """,
                    (table_id, e["min"], e["max"], e["weight"], e["result"], i),
                )

        async def build_type_entries_from_sheet(region_id: Optional[int]) -> List[Dict[str, Any]]:
            tab = _enc_type_tab(region_id)
            ws = wb[tab]
            header_map, data = _read_sheet_rows(ws)
            mode, _ = _detect_mode(tab, header_map, data)

            out: List[Dict[str, Any]] = []
            for _, row in data:
                t = _cell_str(row[header_map["type"]])
                if not t:
                    continue
                if mode == ROLL_RANGE:
                    mi = _to_int(row[header_map["min"]])
                    ma = _to_int(row[header_map["max"]])
                    if mi is None or ma is None:
                        continue
                    out.append({"min": mi, "max": ma, "weight": None, "result": t})
                elif mode == ROLL_WEIGHT:
                    w = _to_int(row[header_map["weight"]])
                    if w is None or w <= 0:
                        continue
                    out.append({"min": None, "max": None, "weight": w, "result": t})
                else:
                    out.append({"min": None, "max": None, "weight": None, "result": t})
            return out

        for region_blob in parsed_regions:
            region_id = region_blob["region_id"]
            enc_type_info = region_blob["enc_type"]
            types: List[str] = enc_type_info["types"]

            # encounter_type table
            type_table_id = await insert_table(
                group_key="encounter_type",
                region_id=region_id,
                type_key=None,
                roll_mode=enc_type_info["mode"],
                max_roll=enc_type_info["max_roll"],
            )
            type_entries = await build_type_entries_from_sheet(region_id)
            await insert_entries(type_table_id, type_entries)

            # typed encounter and reward tables
            for t in types:
                enc_t = region_blob["enc_tables"][t]
                enc_id = await insert_table(
                    group_key="encounter",
                    region_id=region_id,
                    type_key=t,
                    roll_mode=enc_t["mode"],
                    max_roll=enc_t["max_roll"],
                )
                await insert_entries(enc_id, enc_t["entries"])

                rew_t = region_blob["rew_tables"][t]
                rew_id = await insert_table(
                    group_key="reward",
                    region_id=region_id,
                    type_key=t,
                    roll_mode=rew_t["mode"],
                    max_roll=rew_t["max_roll"],
                )
                await insert_entries(rew_id, rew_t["entries"])

        await db.commit()

        counts = ImportCounts(
            encounter_types=total_enc_types,
            encounter_entries=total_enc_entries,
            reward_types=total_enc_types,
            reward_entries=total_rew_entries,
            regions=len(regions) if regional_mode else 0,
        )
        return counts, []

    except Exception as e:
        await db.rollback()
        return None, [ImportError("DB", None, f"Import failed during database write: {e}")]
